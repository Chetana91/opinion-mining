import sys
from nltk.classify import NaiveBayesClassifier
import nltk.metrics
import openpyxl
import collections
from clean_data import clean_training_data
from clean_data import clean_test_data

word_features = None

def main():
	global word_features
	sheet_name = sys.argv[1]

	class_map = {
	0:"neutral",
	1:"positive",
	-1:"negative"
	}
	data_train = openpyxl.load_workbook("training-Obama-Romney-tweets.xlsx")
	train_sheet = data_train.get_sheet_by_name(sheet_name)

	train_tweets = []

	total = train_sheet.get_highest_row()
	num_folds = 10
	offset = 3 # data starts from 3rd row

	for r in xrange(offset,total):
		try:
			if unicode(train_sheet.cell(row=r,column=5).value) is not None:
				tweet_string = unicode(train_sheet.cell(row=r,column=4).value)
				tweet_class_str = unicode(train_sheet.cell(row=r,column=5).value)
				tweet_class = class_map[int(tweet_class_str.strip())]	
				if tweet_string:
					train_tweets.append((tweet_string, tweet_class))
		except Exception:
			pass
	train_tweets = clean_training_data(train_tweets)

	word_features = return_tweet_features(return_tweet_words(train_tweets))
	training_set = nltk.classify.apply_features(obtain_document_features, train_tweets)
	classifier = nltk.NaiveBayesClassifier.train(training_set)


	print "********Classifier Built********"

	test_data_workbook = openpyxl.load_workbook("testing-Obama-Romney-tweets-3labels.xlsx")
	test_sheet = test_data_workbook.get_sheet_by_name(sheet_name)

	test_tweets = []
	row_count = test_sheet.get_highest_row()

	for r in xrange(1,row_count):
		try:
			original_tweet_string = unicode(test_sheet.cell(row=r,column=1).value)
			class_label_str = unicode(test_sheet.cell(row=r,column=2).value)
			class_label = class_map[int(class_label_str.strip())]
			tweet_string = clean_test_data([original_tweet_string])[0]
			test_tweets.append((tweet_string, class_label))
		except Exception as ex:
			pass

	refsets = collections.defaultdict(set)
	testsets = collections.defaultdict(set)

	k=0
	accu = 0
	accu_count=0
	avg_accuracy = 0
	for (tweet, class_label) in test_tweets:
		accu_count+=1
		refsets[class_label].add(k)
		observed = classifier.classify(obtain_document_features(tweet))
		testsets[observed].add((k))
		if observed == class_label:
			accu+=1
		k += 1

	avg_accuracy = accu / float(accu_count)


	print "********Tested********"

	print "Average Accuracy for ",sheet_name," is " , avg_accuracy

	print 'positive precision:', nltk.metrics.precision(refsets['positive'], testsets['positive'])
	print 'pos recall:', nltk.metrics.recall(refsets['positive'], testsets['positive'])
	print '  --- pos F-measure:', nltk.metrics.f_measure(refsets['positive'], testsets['positive'])

	print 'negative precision:', nltk.metrics.precision(refsets['negative'], testsets['negative'])
	print 'neg recall:', nltk.metrics.recall(refsets['negative'], testsets['negative'])
	print '  --- neg F-measure:', nltk.metrics.f_measure(refsets['negative'], testsets['negative'])

	print 'neutral precision:', nltk.metrics.precision(refsets['neutral'], testsets['neutral'])
	print 'new recall:', nltk.metrics.recall(refsets['neutral'], testsets['neutral'])
	print '  --- new F-measure:', nltk.metrics.f_measure(refsets['neutral'], testsets['neutral'])	

def return_tweet_words(train_tweets):
    all_words = []
    for (words, sentiment) in train_tweets:
      all_words.extend(words)
    return all_words


def return_tweet_features(wordlist):
	global word_features
	wordlist = nltk.FreqDist(wordlist)
	word_features = wordlist.keys()
	return word_features

def obtain_document_features(document):
	global word_features
	document_words = set(document)
	features = {}
	for word in word_features:
		features['contains(%s)' % word] = (word in document_words)
	return features

if __name__ == '__main__':
	main()